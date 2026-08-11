from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
import unicodedata
import uuid
from zoneinfo import ZoneInfo

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, or_, select

from ..autenticacion.decoradores import roles
from ..autenticacion.permisos import ROLES_ADMINISTRACION_INSTITUCIONAL
from ..esquemas import error
from ..extensiones import db
from ..modelos import (
    AdminProfile,
    Audit,
    Category,
    Exhibitor,
    Fair,
    FeriaStatus,
    Product,
    ProductStatus,
    ProductiveSector,
    ProductiveUnit,
    ProductiveUnitStatus,
    RegistrationRequest,
    RegistrationRequestSector,
    RegistrationStatus,
    Role,
    SectorStatus,
    UnitSector,
    User,
    UserStatus,
)
from ..servicios import audit

report_bp = Blueprint("reports", __name__)
BOLIVIA_TZ = ZoneInfo("America/La_Paz")
MAX_REPORT_ROWS = 10_000
REPORT_LOGO_PATH = (
    Path(__file__).resolve().parents[3]
    / "frontend"
    / "public"
    / "escudo-reportes.png"
)

STATUS_LABELS = {
    "ACTIVE": "Activo",
    "INACTIVE": "Inactivo",
    "LOCKED": "Bloqueado",
    "BLOCKED": "Bloqueado",
    "PENDING": "Pendiente",
    "APPROVED": "Aprobada",
    "REJECTED": "Rechazada",
    "AVAILABLE": "Disponible",
    "OUT_OF_STOCK": "Agotado",
    "RETIRED": "Retirado",
    "DELETED": "Eliminado",
    "DRAFT": "En preparación",
    "PUBLISHED": "Publicada",
    "FINISHED": "Finalizada",
    "DISABLED": "Cancelada",
}
ROLE_LABELS = {"ADMIN": "Administrador"}
AUDIT_ACTION_LABELS = {
    "CREAR": "Creación",
    "EDITAR": "Edición",
    "ELIMINAR": "Eliminación",
    "RESTAURAR": "Restauración",
    "CAMBIAR_ESTADO": "Cambio de estado",
    "SINCRONIZAR_ESTADO": "Actualización automática del estado",
    "AGREGAR_IMAGEN": "Adición de imagen",
    "EDITAR_IMAGEN": "Edición de imagen",
    "ELIMINAR_IMAGEN": "Eliminación de imagen",
    "ASIGNAR": "Asignación",
    "AUTORIZAR": "Autorización",
    "REVOCAR": "Retiro de autorización",
    "BLOQUEAR": "Bloqueo de cuenta",
    "DESBLOQUEAR": "Desbloqueo de cuenta",
    "INICIAR_SESION": "Inicio de sesión",
    "CERRAR_SESION": "Cierre de sesión",
    "REAUTENTICAR": "Desbloqueo de sesión",
    "RENOVAR_SESION": "Renovación de sesión",
    "SESION_CADUCADA": "Sesión finalizada",
    "INTENTO_FALLIDO": "Intento de acceso fallido",
    "CAMBIAR_CONTRASENA": "Cambio de contraseña",
    "RESTABLECER_CONTRASENA": "Restablecimiento de contraseña",
    "CREAR_SOLICITUD": "Creación de solicitud",
    "APROBAR_SOLICITUD": "Aprobación de solicitud",
    "RECHAZAR_SOLICITUD": "Rechazo de solicitud",
    "ENVIAR_CREDENCIALES": "Envío de datos de acceso",
    "REENVIAR_CREDENCIALES": "Reenvío de datos de acceso",
    "ENVIAR_RECHAZO": "Envío de aviso de rechazo",
    "ENVIAR_RECUPERACION": "Envío de recuperación de contraseña",
    "INTENTO_RECUPERACION_FALLIDO": "Intento de recuperación fallido",
    "GENERAR_REPORTE": "Generación de reporte",
    "CREAR_UNIDAD_PRODUCTIVA": "Registro de unidad productiva",
}
AUDIT_ENTITY_LABELS = {
    "RegistrationRequest": "Solicitud de registro",
    "ProductiveUnit": "Unidad productiva",
    "ProductiveSector": "Sector productivo",
    "FairParticipation": "Participación en feria",
    "FeriaExpositor": "Participación en feria",
    "Fair": "Feria o evento",
    "Product": "Producto",
    "Usuario": "Usuario",
    "Perfil": "Perfil",
    "Unidad": "Unidad administrativa",
    "Categoria": "Categoría",
    "Producto": "Producto",
    "Feria": "Feria",
    "Expositor": "Unidad productiva",
    "Reporte": "Reporte",
}
AUDIT_RESULT_LABELS = {
    "SUCCESS": "Realizada correctamente",
    "PENDING": "Pendiente",
    "FAILED": "No realizada",
}

REPORT_COLUMNS = {
    "solicitudes": {
        "nombre_comercial": "Unidad productiva",
        "razon_social": "Razón social",
        "representante": "Representante",
        "departamento": "Departamento",
        "sectores": "Sectores",
        "nit": "NIT",
        "seprec": "Registro SEPREC",
        "pro_bolivia": "Registro PRO-BOLIVIA",
        "redes_sociales": "Redes sociales",
        "correo": "Correo",
        "telefono": "Teléfono",
        "estado": "Estado",
        "fecha_solicitud": "Fecha de solicitud",
        "fecha_revision": "Fecha de revisión",
    },
    "unidades_productivas": {
        "nombre_comercial": "Unidad productiva",
        "razon_social": "Razón social",
        "representante": "Representante",
        "departamento": "Departamento",
        "sectores": "Sectores",
        "nit": "NIT",
        "seprec": "Registro SEPREC",
        "pro_bolivia": "Registro PRO-BOLIVIA",
        "redes_sociales": "Redes sociales",
        "correo": "Correo",
        "telefono": "Teléfono",
        "estado": "Estado",
        "fecha_aprobacion": "Fecha de aprobación",
    },
    "sectores_productivos": {
        "nombre": "Sector productivo",
        "descripcion": "Descripción",
        "estado": "Estado",
        "unidades": "Unidades productivas relacionadas",
        "creado": "Fecha de registro",
    },
    "productos": {
        "nombre": "Producto",
        "unidad": "Unidad productiva",
        "precio": "Precio de referencia (Bs)",
        "estado": "Estado",
        "descripcion": "Descripción del producto",
        "creado": "Fecha de registro",
    },
    "ferias": {
        "nombre": "Feria",
        "lugar": "Lugar",
        "departamento": "Departamento",
        "fecha_inicio": "Fecha de inicio",
        "fecha_fin": "Fecha final",
        "estado": "Estado",
        "visible": "Visible para el público",
        "descripcion": "Descripción",
        "creado": "Fecha de registro",
    },
    "administradores": {
        "nombre": "Nombre completo",
        "usuario": "Nombre de usuario",
        "correo": "Correo",
        "celular": "Celular",
        "cargo": "Cargo",
        "unidad": "Unidad administrativa",
        "estado": "Estado",
        "creado": "Fecha de registro",
    },
    "auditoria": {
        "fecha": "Fecha y hora",
        "usuario": "Nombre de usuario",
        "accion": "Actividad realizada",
        "entidad": "Tipo de información",
        "descripcion": "Detalle",
        "resultado": "Estado de la actividad",
    },
    # Compatibilidad con enlaces de reportes de la version anterior.
    "categorias": {
        "nombre": "Categoría",
        "descripcion": "Descripción",
        "estado": "Estado",
        "creado": "Fecha de registro",
        "actualizado": "Última actualización",
    },
}

REPORT_TITLES = {
    "solicitudes": "Solicitudes de registro",
    "unidades_productivas": "Unidades productivas",
    "sectores_productivos": "Sectores productivos",
    "productos": "Productos",
    "ferias": "Ferias",
    "administradores": "Administradores",
    "auditoria": "Auditoría",
    "categorias": "Categorías",
}
GENERAL_RESOURCES = [
    "solicitudes",
    "unidades_productivas",
    "sectores_productivos",
    "productos",
    "ferias",
    "administradores",
    "auditoria",
]


def enum_value(value):
    return value.value if hasattr(value, "value") else value


def display(value):
    value = enum_value(value)
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, datetime):
        local = value.astimezone(BOLIVIA_TZ) if value.tzinfo else value
        return local.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if value in STATUS_LABELS:
        return STATUS_LABELS[value]
    if value in ROLE_LABELS:
        return ROLE_LABELS[value]
    return str(value)


def safe_filename_label(value):
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-zA-Z0-9]+", "_", normalized).strip("_")


def parse_date(name):
    raw = request.args.get(name, "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError("Ingrese una fecha válida") from exc


def parse_list(name):
    """Acepta parámetros repetidos y valores separados por comas."""
    values = []
    for raw in request.args.getlist(name):
        values.extend(item.strip() for item in raw.split(","))
    return list(dict.fromkeys(item for item in values if item))


def parse_uuids(name):
    values = parse_list(name)
    try:
        return [uuid.UUID(item) for item in values]
    except ValueError as exc:
        raise ValueError("Una de las opciones seleccionadas no es válida") from exc


def parse_optional_bool(name):
    raw = request.args.get(name, "").strip().lower()
    if not raw:
        return None
    if raw not in {"true", "false"}:
        raise ValueError("Seleccione una opción válida")
    return raw == "true"


def parse_multi_values(name):
    values = []
    for raw in request.args.getlist(name):
        for item in raw.split(","):
            cleaned = item.strip()
            if cleaned:
                values.append(cleaned)
    return list(dict.fromkeys(values))


def apply_created_dates(query, column):
    start, end = parse_date("date_from"), parse_date("date_to")
    if start:
        query = query.where(func.date(column) >= start)
    if end:
        query = query.where(func.date(column) <= end)
    return query


def apply_presence_filter(query, column, name):
    present = parse_optional_bool(name)
    if present is True:
        return query.where(column.is_not(None), func.length(func.trim(column)) > 0)
    if present is False:
        return query.where(or_(column.is_(None), func.length(func.trim(column)) == 0))
    return query


def has_social_networks(model):
    return or_(model.facebook_url.is_not(None), model.instagram_url.is_not(None), model.tiktok_url.is_not(None))


def apply_social_filter(query, model):
    present = parse_optional_bool("has_social")
    condition = has_social_networks(model)
    if present is True:
        return query.where(condition)
    if present is False:
        return query.where(~condition)
    return query


def selected_columns(resource):
    available = REPORT_COLUMNS[resource]
    requested = [item.strip() for item in request.args.get("columns", "").split(",") if item.strip()]
    if not requested:
        return list(available)
    chosen = [item for item in requested if item in available]
    if not chosen:
        raise ValueError("Seleccione al menos una columna valida")
    return chosen


def sector_names(link_model, owner_field, owner_id):
    names = db.session.scalars(
        select(ProductiveSector.nombre)
        .join(link_model, link_model.productive_sector_id == ProductiveSector.id)
        .where(getattr(link_model, owner_field) == owner_id)
        .order_by(ProductiveSector.nombre)
    ).all()
    return ", ".join(names)


def apply_organization_filters(query, model, link_model, owner_field):
    term = request.args.get("q", "").strip()
    if term:
        pattern = f"%{term}%"
        query = query.where(or_(model.nombre_comercial.ilike(pattern), model.razon_social.ilike(pattern)))
    status = request.args.get("status", "")
    if status:
        enum_class = RegistrationStatus if model is RegistrationRequest else ProductiveUnitStatus
        try:
            query = query.where(model.estado == enum_class(status))
        except ValueError as exc:
            raise ValueError("Seleccione un estado válido") from exc
    departments = parse_list("departments") or parse_list("department")
    if departments:
        query = query.where(model.departamento.in_(departments))
    sector_ids = parse_uuids("sector_ids") or parse_uuids("sector_id")
    if sector_ids:
        query = query.where(
            select(link_model.id)
            .where(
                getattr(link_model, owner_field) == model.id,
                link_model.productive_sector_id.in_(sector_ids),
            )
            .exists()
        )
    query = apply_presence_filter(query, model.nit, "has_nit")
    query = apply_presence_filter(query, model.registro_seprec, "has_seprec")
    query = apply_presence_filter(query, model.registro_pro_bolivia, "has_pro_bolivia")
    return apply_social_filter(query, model)


def organization_row(item, link_model, owner_field):
    networks = [name for name, value in (("Facebook", item.facebook_url), ("Instagram", item.instagram_url), ("TikTok", item.tiktok_url)) if value]
    return {
        "nombre_comercial": item.nombre_comercial,
        "razon_social": item.razon_social,
        "representante": item.nombre_representante,
        "departamento": item.departamento,
        "sectores": sector_names(link_model, owner_field, item.id),
        "nit": item.nit,
        "seprec": item.registro_seprec,
        "pro_bolivia": item.registro_pro_bolivia,
        "redes_sociales": ", ".join(networks) if networks else "No",
        "correo": item.correo_electronico,
        "telefono": item.telefono_whatsapp,
        "estado": item.estado,
    }


def registration_rows():
    query = apply_organization_filters(
        select(RegistrationRequest), RegistrationRequest, RegistrationRequestSector, "registration_request_id"
    )
    query = apply_created_dates(query, RegistrationRequest.created_at).order_by(RegistrationRequest.created_at.desc())
    rows = []
    for item in db.session.scalars(query.limit(MAX_REPORT_ROWS)):
        row = organization_row(item, RegistrationRequestSector, "registration_request_id")
        row.update(fecha_solicitud=item.created_at, fecha_revision=item.fecha_revision)
        rows.append(row)
    return rows


def productive_unit_rows():
    query = apply_organization_filters(
        select(ProductiveUnit), ProductiveUnit, UnitSector, "productive_unit_id"
    )
    query = query.where(ProductiveUnit.deleted_at.is_(None))
    query = apply_created_dates(query, ProductiveUnit.fecha_aprobacion).order_by(ProductiveUnit.fecha_aprobacion.desc())
    rows = []
    for item in db.session.scalars(query.limit(MAX_REPORT_ROWS)):
        row = organization_row(item, UnitSector, "productive_unit_id")
        row["fecha_aprobacion"] = item.fecha_aprobacion
        rows.append(row)
    return rows


def productive_sector_rows():
    query = select(ProductiveSector).where(ProductiveSector.deleted_at.is_(None))
    term = request.args.get("q", "").strip()
    if term:
        query = query.where(ProductiveSector.nombre.ilike(f"%{term}%"))
    status = request.args.get("status", "")
    if status:
        try:
            query = query.where(ProductiveSector.estado == SectorStatus(status))
        except ValueError as exc:
            raise ValueError("Seleccione un estado válido") from exc
    query = apply_created_dates(query, ProductiveSector.created_at).order_by(ProductiveSector.nombre)
    rows = []
    for item in db.session.scalars(query.limit(MAX_REPORT_ROWS)):
        count = db.session.scalar(
            select(func.count(UnitSector.id)).where(
                UnitSector.productive_sector_id == item.id,
                UnitSector.estado == SectorStatus.ACTIVE,
            )
        )
        rows.append({"nombre": item.nombre, "descripcion": item.descripcion, "estado": item.estado, "unidades": count, "creado": item.created_at})
    return rows


def product_rows():
    query = (
        select(Product, ProductiveUnit)
        .join(ProductiveUnit, Product.productive_unit_id == ProductiveUnit.id)
        .where(Product.productive_unit_id.is_not(None), Product.deleted_at.is_(None))
    )
    term = request.args.get("q", "").strip()
    if term:
        query = query.where(or_(Product.nombre_comercial.ilike(f"%{term}%"), ProductiveUnit.nombre_comercial.ilike(f"%{term}%")))
    unit_ids = parse_uuids("productive_unit_ids") or parse_uuids("productive_unit_id")
    if unit_ids:
        query = query.where(Product.productive_unit_id.in_(unit_ids))
    status = request.args.get("status", "")
    if status:
        try:
            query = query.where(Product.estado == ProductStatus(status))
        except ValueError as exc:
            raise ValueError("Seleccione un estado válido") from exc
    try:
        price_min = float(request.args["price_min"]) if request.args.get("price_min") else None
        price_max = float(request.args["price_max"]) if request.args.get("price_max") else None
    except ValueError as exc:
        raise ValueError("Ingrese precios válidos") from exc
    if (price_min is not None and price_min < 0) or (price_max is not None and price_max < 0):
        raise ValueError("Los precios no pueden ser negativos")
    if price_min is not None and price_max is not None and price_min > price_max:
        raise ValueError("El precio mínimo no puede ser mayor que el precio máximo")
    if price_min is not None:
        query = query.where(Product.precio_referencia >= price_min)
    if price_max is not None:
        query = query.where(Product.precio_referencia <= price_max)
    query = apply_created_dates(query, Product.created_at).order_by(Product.created_at.desc())
    return [{
        "nombre": product.nombre_comercial or product.nombre,
        "unidad": unit.nombre_comercial,
        "precio": product.precio_referencia,
        "estado": product.estado,
        "descripcion": product.descripcion_tecnica or product.descripcion,
        "creado": product.created_at,
    } for product, unit in db.session.execute(query.limit(MAX_REPORT_ROWS))]


def fair_rows():
    query = select(Fair).where(Fair.deleted_at.is_(None))
    term = request.args.get("q", "").strip()
    if term:
        pattern = f"%{term}%"
        query = query.where(or_(Fair.nombre.ilike(pattern), Fair.lugar.ilike(pattern), Fair.ubicacion.ilike(pattern)))
    location = request.args.get("location", "").strip()
    if location:
        query = query.where(or_(Fair.lugar == location, Fair.ubicacion == location))
    status = request.args.get("status", "")
    if status:
        try:
            query = query.where(Fair.estado == FeriaStatus(status))
        except ValueError as exc:
            raise ValueError("Seleccione un estado válido") from exc
    start, end = parse_date("date_from"), parse_date("date_to")
    if start and end and end <= start:
        raise ValueError("La fecha final debe ser posterior a la fecha inicial")
    if start:
        query = query.where(Fair.fecha_fin >= start)
    if end:
        query = query.where(Fair.fecha_inicio <= end)
    query = query.order_by(Fair.fecha_inicio.desc())
    return [{
        "nombre": item.nombre,
        "lugar": item.ubicacion or item.lugar,
        "departamento": item.departamento,
        "fecha_inicio": item.fecha_inicio,
        "fecha_fin": item.fecha_fin,
        "estado": item.estado,
        "visible": item.visible_publicamente,
        "descripcion": item.descripcion,
        "creado": item.created_at,
    } for item in db.session.scalars(query.limit(MAX_REPORT_ROWS))]


def administrator_rows():
    query = select(User, AdminProfile).outerjoin(AdminProfile).where(User.role == Role.ADMIN, User.deleted_at.is_(None))
    term = request.args.get("q", "").strip()
    if term:
        pattern = f"%{term}%"
        query = query.where(or_(User.first_name.ilike(pattern), User.last_name.ilike(pattern), User.username.ilike(pattern), User.email.ilike(pattern)))
    status = request.args.get("status", "")
    if status:
        try:
            query = query.where(User.status == UserStatus(status))
        except ValueError as exc:
            raise ValueError("Seleccione un estado válido") from exc
    query = apply_created_dates(query, User.created_at).order_by(User.created_at.desc())
    return [{
        "nombre": " ".join(filter(None, [user.first_name, user.apellido_paterno or user.last_name, user.apellido_materno])),
        "usuario": user.username,
        "correo": user.email,
        "celular": user.phone,
        "cargo": profile.cargo if profile else None,
        "unidad": profile.unidad if profile else None,
        "estado": user.status,
        "creado": user.created_at,
    } for user, profile in db.session.execute(query.limit(MAX_REPORT_ROWS))]


def public_audit_description(value):
    if not value:
        return "Actividad registrada"
    text = str(value)
    replacements = {
        "RegistrationRequest": "solicitud de registro",
        "ProductiveUnit": "unidad productiva",
        "ProductiveSector": "sector productivo",
        "FairParticipation": "participación en feria",
        "FeriaExpositor": "participación en feria",
        "unidades_productivas": "unidades productivas",
        "sectores_productivos": "sectores productivos",
        "solicitudes": "solicitudes de registro",
        "administradores": "administradores",
        "auditoria": "auditoría",
        "xlsx": "Excel",
    }
    for source, label in replacements.items():
        text = re.sub(rf"\b{re.escape(source)}\b", label, text, flags=re.IGNORECASE)
    return text.replace("_", " ")


def audit_rows():
    query = select(Audit, User).outerjoin(User, Audit.user_id == User.id)
    term = request.args.get("q", "").strip()
    if term:
        pattern = f"%{term}%"
        query = query.where(or_(Audit.accion.ilike(pattern), Audit.descripcion.ilike(pattern), User.username.ilike(pattern)))
    actions = parse_list("actions") or parse_list("action")
    if actions:
        query = query.where(Audit.accion.in_(actions))
    query = apply_created_dates(query, Audit.created_at).order_by(Audit.created_at.desc())
    return [{
        "fecha": item.created_at,
        "usuario": user.username if user else "Sistema",
        "accion": AUDIT_ACTION_LABELS.get(item.accion, item.accion.replace("_", " ").capitalize()),
        "entidad": AUDIT_ENTITY_LABELS.get(item.entidad, item.entidad.replace("_", " ").capitalize()),
        "descripcion": public_audit_description(item.descripcion),
        "resultado": AUDIT_RESULT_LABELS.get(item.resultado, item.resultado.replace("_", " ").capitalize()),
    } for item, user in db.session.execute(query.limit(MAX_REPORT_ROWS))]


def category_rows():
    query = select(Category).where(Category.deleted_at.is_(None))
    if request.args.get("status") in {"active", "inactive"}:
        query = query.where(Category.estado.is_(request.args["status"] == "active"))
    query = apply_created_dates(query, Category.created_at).order_by(Category.nombre)
    return [{"nombre": item.nombre, "descripcion": item.descripcion, "estado": item.estado, "creado": item.created_at, "actualizado": item.updated_at} for item in db.session.scalars(query.limit(MAX_REPORT_ROWS))]


ROW_BUILDERS = {
    "solicitudes": registration_rows,
    "unidades_productivas": productive_unit_rows,
    "sectores_productivos": productive_sector_rows,
    "productos": product_rows,
    "ferias": fair_rows,
    "administradores": administrator_rows,
    "auditoria": audit_rows,
    "categorias": category_rows,
}


def report_sections(resource):
    resources = GENERAL_RESOURCES if resource == "general" else [resource]
    return [(item, selected_columns(item) if resource != "general" else list(REPORT_COLUMNS[item]), ROW_BUILDERS[item]()) for item in resources]


def build_xlsx(sections, generated_at):
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1B7340")
    header_font = Font(color="FFFFFF", bold=True)
    for resource, columns, rows in sections:
        sheet = workbook.create_sheet(REPORT_TITLES[resource][:31])
        headers = [REPORT_COLUMNS[resource][column] for column in columns]
        if resource == "categorias":
            sheet.append(headers)
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in rows:
                sheet.append([display(row.get(column)) for column in columns])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            continue
        column_count = max(1, len(headers))
        if REPORT_LOGO_PATH.exists():
            logo = ExcelImage(str(REPORT_LOGO_PATH))
            logo.width = 66
            logo.height = 66
            center_column = max(1, (column_count + 1) // 2)
            sheet.add_image(logo, f"{get_column_letter(center_column)}1")
        for row_number in range(1, 4):
            sheet.row_dimensions[row_number].height = 18
        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=column_count)
        title = sheet.cell(4, 1, "Ministerio de\nDesarrollo productivo\nRural y agua")
        title.font = Font(bold=False, size=12, color="4A4A4A")
        title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[4].height = 44
        sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=column_count)
        subtitle = sheet.cell(5, 1, f"Reporte de {REPORT_TITLES[resource]} - Fecha de emisión: {generated_at.strftime('%d/%m/%Y %H:%M:%S')}")
        subtitle.alignment = Alignment(horizontal="center")
        for index, header in enumerate(headers, 1):
            sheet.cell(7, index, header)
        for cell in sheet[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            values = [display(row.get(column)) for column in columns]
            sheet.append([f"'{value}" if isinstance(value, str) and value.startswith(("=", "+", "-", "@")) else value for value in values])
        sheet.freeze_panes = "A8"
        sheet.auto_filter.ref = f"A7:{get_column_letter(len(headers))}{max(7, sheet.max_row)}"
        sheet.sheet_view.showGridLines = False
        for index, header in enumerate(headers, 1):
            values = [str(sheet.cell(row=row, column=index).value or "") for row in range(7, min(sheet.max_row, 200) + 1)]
            sheet.column_dimensions[get_column_letter(index)].width = min(45, max(12, len(header) + 2, *(len(value) + 2 for value in values)))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _institutional_page(canvas, document):
    width, height = landscape(A4)
    canvas.saveState()
    if REPORT_LOGO_PATH.exists():
        logo_size = 24 * mm
        canvas.drawImage(
            str(REPORT_LOGO_PATH),
            width / 2 - logo_size / 2,
            height - 31 * mm,
            logo_size,
            logo_size,
            preserveAspectRatio=True,
            mask="auto",
            anchor="c",
        )
    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(colors.HexColor("#555555"))
    canvas.drawCentredString(width / 2, height - 34 * mm, "Ministerio de")
    canvas.drawCentredString(width / 2, height - 38 * mm, "Desarrollo productivo")
    canvas.drawCentredString(width / 2, height - 42 * mm, "Rural y agua")
    canvas.setStrokeColor(colors.HexColor("#C6C6C6"))
    canvas.line(10 * mm, 25 * mm, width - 10 * mm, 25 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#555555"))
    canvas.drawRightString(width - 13 * mm, 19 * mm, "Ministerio de Desarrollo Productivo, Rural y Agua")
    canvas.setFont("Helvetica", 7.5)
    canvas.drawRightString(width - 13 * mm, 14.5 * mm, "Av. Mariscal Santa Cruz, Edif. Centro de Comunicaciones La Paz")
    canvas.drawRightString(width - 13 * mm, 10.5 * mm, "Piso 20 - Teléfonos: +591 (2) 2184444 - Fax: (591-2) 2124933")
    bar_y, bar_h, bar_x, segment = 7 * mm, 2.5 * mm, 10 * mm, 32 * mm
    for index, color in enumerate(("#D71920", "#F3DC19", "#148447")):
        canvas.setFillColor(colors.HexColor(color))
        canvas.rect(bar_x + index * segment, bar_y, segment, bar_h, stroke=0, fill=1)
    canvas.restoreState()


def build_pdf(sections, generated_at):
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=47 * mm,
        bottomMargin=30 * mm,
        title="Reporte de Ferias Productivas Bolivia",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#1B7340"), fontSize=16)
    meta_style = ParagraphStyle("ReportMeta", parent=styles["Normal"], alignment=TA_CENTER, fontSize=8, textColor=colors.HexColor("#555555"))
    cell_style = ParagraphStyle("ReportCell", parent=styles["BodyText"], fontSize=6.2, leading=7.5)
    header_style = ParagraphStyle("ReportHeader", parent=cell_style, textColor=colors.white, alignment=TA_CENTER)
    story = []
    for section_index, (resource, columns, rows) in enumerate(sections):
        if section_index:
            story.append(PageBreak())
        story.append(Paragraph(f"Reporte de {REPORT_TITLES[resource]}", title_style))
        story.append(Paragraph(f"Fecha de emisión: {generated_at.strftime('%d/%m/%Y %H:%M:%S')} | Registros: {len(rows)}", meta_style))
        story.append(Spacer(1, 4 * mm))
        data = [[Paragraph(REPORT_COLUMNS[resource][column], header_style) for column in columns]]
        for row in rows:
            data.append([Paragraph(display(row.get(column)).replace("&", "&amp;").replace("<", "&lt;"), cell_style) for column in columns])
        if len(data) == 1:
            data.append([Paragraph("Sin registros para los filtros seleccionados", cell_style)] + [""] * (len(columns) - 1))
        available_width = landscape(A4)[0] - 20 * mm
        table = Table(data, repeatRows=1, colWidths=[available_width / len(columns)] * len(columns))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B7340")),
            ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#CBD5DF")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F3")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
    document.build(story, onFirstPage=_institutional_page, onLaterPages=_institutional_page)
    output.seek(0)
    return output


@report_bp.get("/reports/options")
@roles(*ROLES_ADMINISTRACION_INSTITUCIONAL)
def report_options():
    actions = db.session.scalars(select(Audit.accion).distinct().order_by(Audit.accion)).all()
    sectors = db.session.execute(select(ProductiveSector.id, ProductiveSector.nombre).where(ProductiveSector.deleted_at.is_(None)).order_by(ProductiveSector.nombre)).all()
    units = db.session.execute(select(ProductiveUnit.id, ProductiveUnit.nombre_comercial).where(ProductiveUnit.deleted_at.is_(None)).order_by(ProductiveUnit.nombre_comercial)).all()
    fair_locations = db.session.scalars(
        select(func.coalesce(Fair.ubicacion, Fair.lugar))
        .where(Fair.deleted_at.is_(None))
        .distinct()
        .order_by(func.coalesce(Fair.ubicacion, Fair.lugar))
    ).all()
    departments = sorted(set(db.session.scalars(select(ProductiveUnit.departamento).distinct()).all()) | set(db.session.scalars(select(RegistrationRequest.departamento).distinct()).all()))
    return {
        "resources": [
            {"value": key, "label": REPORT_TITLES[key], "columns": [{"value": name, "label": label} for name, label in REPORT_COLUMNS[key].items()]}
            for key in ROW_BUILDERS
        ],
        "actions": actions,
        "sectors": [{"value": str(item.id), "label": item.nombre} for item in sectors],
        "productive_units": [{"value": str(item.id), "label": item.nombre_comercial} for item in units],
        "fair_locations": [{"value": item, "label": item} for item in fair_locations if item],
        "departments": [item for item in departments if item],
    }


@report_bp.get("/reports/<resource>")
@roles(*ROLES_ADMINISTRACION_INSTITUCIONAL)
def download_report(resource):
    if resource not in {*ROW_BUILDERS, "general"}:
        return error("Tipo de reporte no válido", 404)
    report_format = request.args.get("format", "pdf").lower()
    if report_format not in {"pdf", "xlsx"}:
        return error("El formato debe ser PDF o Excel")
    try:
        sections = report_sections(resource)
    except ValueError as exc:
        return error(str(exc))
    generated_at = datetime.now(BOLIVIA_TZ)
    output = build_pdf(sections, generated_at) if report_format == "pdf" else build_xlsx(sections, generated_at)
    extension = "pdf" if report_format == "pdf" else "xlsx"
    report_label = "general" if resource == "general" else REPORT_TITLES[resource]
    filename = f"Reporte_{safe_filename_label(report_label)}_{generated_at.strftime('%Y-%m-%d_%H-%M-%S')}.{extension}"
    total = sum(len(rows) for _, _, rows in sections)
    audit("GENERAR_REPORTE", "Reporte", description=f"Reporte {resource} generado en {extension.upper()} con {total} registros")
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf" if extension == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
